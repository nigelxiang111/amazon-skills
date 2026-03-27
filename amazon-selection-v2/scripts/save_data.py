# 数据保存器 v1.0 — Amazon 选品技能包
# 将选品分析数据保存为多 Sheet Excel 文件，附带数据来源和时间戳
#
# 依赖优先级：
#   pandas + openpyxl（pip install pandas openpyxl）→ 最佳体验
#   openpyxl（pip install openpyxl）→ 基础 Excel
#   内置 csv 模块 → 降级为 CSV 目录
#
# 使用方法：
#   1. Claude 将各模块采集到的真实数据填入下方变量
#   2. 修改 TASK_NAME 为本次分析对象（如"硅胶冰块托盘"）
#   3. 将本脚本写入 /tmp/save_data.py 并执行 python3 /tmp/save_data.py
#   4. 文件保存路径：桌面（Windows/Mac 均自动适配）
#      命名格式：{TASK_NAME}分析_{YYYYMMDD}_{HHMMSS}.xlsx

import datetime
import os
import sys

# ═══════════════════════════════════════
# ↓ 由 Claude 填入真实采集到的数据
# ═══════════════════════════════════════

TASK_NAME = "选品分析"       # 分析对象名称，用于文件名前缀
AMZ_SITE  = "US"             # 亚马逊站点

# ─── Sheet 1：市场数据 ───────────────────────────────────────────────────────
# 每行：[指标名称, 数值, 数据来源]  来源填"卖家精灵"或"sorftime"
market_rows = [
    # 示例（Claude 替换为真实数据）：
    # ["月销量(件)", "12万", "卖家精灵"],
    # ["均价($)",    "29.9",  "卖家精灵"],
    # ["Top3垄断",   "38%",   "卖家精灵"],
    # ["亚马逊自营", "5%",    "卖家精灵"],
    # ["新品占比",   "22%",   "卖家精灵"],
    # ["生命周期",   "成长期", "综合判断"],
]

# ─── Sheet 2：产品/竞品数据 ──────────────────────────────────────────────────
# 每行：[ASIN, 产品名, 月销量, 价格($), 评分, 评论数, BSR, 数据来源]
product_rows = [
    # 示例：
    # ["B0CKRQCHRS", "Silicone Ice Cube Tray", 3200, 12.99, 4.5, 8900, 312, "卖家精灵"],
]

# ─── Sheet 3：关键词数据 ─────────────────────────────────────────────────────
# 每行：[关键词, 月搜索量, CPC($), 竞争度, 趋势, 数据来源]
keyword_rows = [
    # 示例：
    # ["silicone ice cube tray", "580000", "0.82", "高", "↑上升", "卖家精灵"],
    # ["large ice cube mold",   "120000", "0.65", "中", "↑上升", "卖家精灵"],
]

# ─── Sheet 4：评论/用户痛点 ──────────────────────────────────────────────────
# 每行：[ASIN, 星级, 摘要, 类型(正面/负面/痛点), 数据来源]
review_rows = [
    # 示例：
    # ["B0CKRQCHRS", "2", "冰块太容易碎，倒的时候粘住", "痛点", "卖家精灵"],
]

# ─── Sheet 5：利润分析 ───────────────────────────────────────────────────────
# 每行：[情景, 售价($), 采购价(¥), FBA($), 佣金($), 广告费($), 净利润($), 毛利率, ROI]
profit_rows = [
    # 示例（来自 pnl.py 输出）：
    # ["乐观", 29.9, 18.0, 4.75, 4.49, 5.23, 9.5, "31.8%", "67%"],
    # ["基准", 29.9, 20.0, 4.75, 4.49, 7.48, 5.5, "18.4%", "34%"],
    # ["悲观", 29.9, 22.0, 4.75, 4.49, 10.47, 0.8,  "2.7%",  "4%"],
]

# ─── Sheet 6：1688货源 ───────────────────────────────────────────────────────
# 每行：[商品名, 价格(¥), 起订量, 供应商, 数据来源]
supply_rows = [
    # 示例：
    # ["硅胶冰格大号", "8.5", "50", "广东某厂", "sorftime/1688"],
]

# ─── Sheet 7：TikTok热度 ─────────────────────────────────────────────────────
# 每行：[产品名, 销量, 价格, 达人数, 热度评级, 数据来源]
tiktok_rows = [
    # 示例：
    # ["silicone ice mold", 1200, 15.9, 34, "中", "sorftime/TikTok"],
]

# ═══════════════════════════════════════
# 输出路径配置
# ═══════════════════════════════════════

def get_desktop_path() -> str:
    """获取桌面路径，兼容 Windows / Mac / Linux"""
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    if not os.path.isdir(desktop):
        desktop = os.path.expanduser("~")
    return desktop

now = datetime.datetime.now()
timestamp = now.strftime("%Y%m%d_%H%M%S")
filename_base = f"{TASK_NAME}分析_{timestamp}"
save_dir = get_desktop_path()

# ═══════════════════════════════════════
# 数据组织
# ═══════════════════════════════════════

sheets_config = [
    {
        "name": "市场数据",
        "headers": ["指标", "数值", "数据来源"],
        "rows": market_rows,
    },
    {
        "name": "产品数据",
        "headers": ["ASIN", "产品名", "月销量(件)", "价格($)", "评分", "评论数", "BSR", "数据来源"],
        "rows": product_rows,
    },
    {
        "name": "关键词数据",
        "headers": ["关键词", "月搜索量", "CPC($)", "竞争度", "趋势", "数据来源"],
        "rows": keyword_rows,
    },
    {
        "name": "评论与痛点",
        "headers": ["ASIN", "星级", "评论摘要", "类型", "数据来源"],
        "rows": review_rows,
    },
    {
        "name": "利润分析",
        "headers": ["情景", "售价($)", "采购价(¥)", "FBA($)", "佣金($)", "广告费($)", "净利润($)", "毛利率", "ROI"],
        "rows": profit_rows,
    },
    {
        "name": "1688货源",
        "headers": ["商品名", "价格(¥)", "起订量", "供应商", "数据来源"],
        "rows": supply_rows,
    },
    {
        "name": "TikTok热度",
        "headers": ["产品名", "销量", "价格", "达人数", "热度评级", "数据来源"],
        "rows": tiktok_rows,
    },
]

meta_rows = [
    ["分析对象",   TASK_NAME],
    ["亚马逊站点", AMZ_SITE],
    ["生成时间",   now.strftime("%Y-%m-%d %H:%M:%S")],
    ["数据来源",   "卖家精灵（优先）+ sorftime（补充）+ 1688（货源）"],
    ["脚本版本",   "save_data.py v1.0"],
]

# ═══════════════════════════════════════
# 保存逻辑（优先 pandas，降级 openpyxl，再降级 CSV）
# ═══════════════════════════════════════

def save_with_pandas(filepath: str) -> None:
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font   = Font(color="FFFFFF", bold=True, size=11)
    meta_fill     = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    thin_border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # 写数据 sheet
        for cfg in sheets_config:
            if not cfg["rows"]:
                continue
            df = pd.DataFrame(cfg["rows"], columns=cfg["headers"])
            df.to_excel(writer, sheet_name=cfg["name"], index=False)

            ws = writer.sheets[cfg["name"]]
            # 美化表头
            for col_idx, _ in enumerate(cfg["headers"], start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill   = header_fill
                cell.font   = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            # 添加边框 + 自动列宽
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
            for col in ws.columns:
                width = max(len(str(c.value or "")) for c in col) + 4
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(width, 40)

        # 写元数据 sheet
        meta_df = pd.DataFrame(meta_rows, columns=["字段", "值"])
        meta_df.to_excel(writer, sheet_name="数据说明", index=False)
        ws_meta = writer.sheets["数据说明"]
        for cell in ws_meta[1]:
            cell.fill   = meta_fill
            cell.font   = Font(bold=True)
            cell.border = thin_border
        for row in ws_meta.iter_rows(min_row=2, max_row=ws_meta.max_row, max_col=2):
            for cell in row:
                cell.border = thin_border


def save_with_openpyxl(filepath: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    wb.remove(wb.active)  # type: ignore

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cfg in sheets_config:
        if not cfg["rows"]:
            continue
        ws = wb.create_sheet(cfg["name"])
        ws.append(cfg["headers"])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in cfg["rows"]:
            ws.append(row)

    ws_meta = wb.create_sheet("数据说明")
    ws_meta.append(["字段", "值"])
    for row in meta_rows:
        ws_meta.append(row)

    wb.save(filepath)


def save_as_csv(base_dir: str) -> str:
    import csv
    csv_dir = os.path.join(base_dir, filename_base)
    os.makedirs(csv_dir, exist_ok=True)

    for cfg in sheets_config:
        if not cfg["rows"]:
            continue
        csv_path = os.path.join(csv_dir, f"{cfg['name']}.csv")
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(cfg["headers"])
            writer.writerows(cfg["rows"])

    meta_path = os.path.join(csv_dir, "数据说明.csv")
    with open(meta_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["字段", "值"])
        writer.writerows(meta_rows)

    return csv_dir


# ─── 主执行逻辑 ────────────────────────────────────────────────────────────────

filepath = os.path.join(save_dir, f"{filename_base}.xlsx")

try:
    import pandas  # noqa: F401
    import openpyxl  # noqa: F401
    save_with_pandas(filepath)
    print(f"✅ Excel 已保存（pandas）：{filepath}")

except ImportError:
    try:
        import openpyxl  # noqa: F401
        save_with_openpyxl(filepath)
        print(f"✅ Excel 已保存（openpyxl）：{filepath}")

    except ImportError:
        out_dir = save_as_csv(save_dir)
        print(f"⚠️  pandas/openpyxl 均不可用，已降级保存为 CSV 目录：{out_dir}")
        print("   建议：pip install pandas openpyxl 以启用 Excel 格式")
        sys.exit(0)

# ─── 输出保存摘要 ──────────────────────────────────────────────────────────────

saved_sheets = [c["name"] for c in sheets_config if c["rows"]]
print(f"   包含 Sheet：{', '.join(saved_sheets) if saved_sheets else '（所有数据为空，仅保存元数据）'}")
print(f"   生成时间：{now.strftime('%Y-%m-%d %H:%M:%S')}")
print(f"   数据来源：卖家精灵 + sorftime + 1688")
